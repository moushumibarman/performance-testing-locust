from locust import User, task
import websocket
import time
import base64
import uuid
import re
import json
import html
from openpyxl import Workbook, load_workbook
import os
from xml.sax.saxutils import escape
import gevent
from gevent.lock import Semaphore
import random
from gevent import monkey

# ----------- GEVENT MONKEY PATCH -----------
monkey.patch_all()

LOG_FILE = "set4responses.xlsx"

# ─────────────────────────────────────────────────────────────
#  BOT DEFINITIONS
#  Each bot has:
#    - username / password / recipient  → unique credentials
#    - queries                          → its own question set
#    - bot_id                           → label used in Excel
# ─────────────────────────────────────────────────────────────
def build_payload(msg, from_jid, task_id, support_key, support_id):
    return json.dumps({
        "key": str(uuid.uuid4().hex),
        "type": "SUPPORT",
        "supportChatMessage": {
            "fromJid": from_jid,
            "senderName": "Guest User",
            "message": msg,
            "taskId": task_id,
            "ticketId": "",
            "enquiryId": 0,
            "leadId": 0,
            "sentDate": int(time.time() * 1000),
            "supportId": support_id,
            "supportKey": support_key,
            "chatbotMessageKey": None,
            "action": None,
            "additionalData": None,
            "productDetail": None,
            "chatbotFlow": None,
            "conversationId": None,
            "isSentByBot": True,
            "traceId": str(uuid.uuid4())
        },
        "traceId": str(uuid.uuid4())
    })

# ── Question sets per bot (customise freely) ──────────────────
BOT_QUESTION_SETS = {
    "bot_1": [

        #  enrollment set 4.1
        "How can I change my account information?",
        "What if I miss a payment or am late on a payment?",
        "How much is the late payment fee?",
        "Will deleting my account also cancel my electricity service?",
        "How can I view my rights as a customer?",
        "What are your terms of service?",
        "I want to enroll",
        "Yes",
        "yes"     
    ],

    "bot_2": [

        #  enrollment set 4.2
        "Who is Champion Energy?",
        "What customers does Champion Energy serve?",
        "Can I get my deposit waived?",
        "I want to enroll",
        "Yes",
        "no" 
    ],

    "bot_3": [

        #  enrollment set 4.3
        "Will deleting my account also cancel my electricity service?",
        "How can I view my rights as a customer?",
        "What are your terms of service?",
        "I want to enroll",
        "no",
        "test address",
        "yes",
        "yes"      
    ],

    # "bot_4": [
    #     # Pay bill [quick pay]
    #     "I need to make a quick payment for account 993871",
    #     "Pay my bill fast using zip 7082",

    #     # Irrelevant chat
    #     "How do I learn to code?",
    #     "What's the latest news?",

    #     # New customer [General FAQ and Enrollment]
    #     "How do I start service at a new address in 7082?",
    #     "I want to switch to Champions Energy, how do I start?",

    #     # Billing [current balance]
    #     "What's my current bill amount for account 993871?",
    #     "What are my current charges for Sanya?",

    #     # Billing [historic balance]
    #     "What did I pay last month for account 993871?",
    #     "Show me old bills for last name Sanya",

    #     # Current rate plan
    #     "What plan is account 993871 on?",
    #     "Current plan information for Sanya",

    #     # Renewal
    #     "Renewal options for account 993871",
    #     "My contract ends soon, how do I renew for Sanya?",

    #     # Contract details
    #     "What's my early termination fee for account 993871?",
    #     "Contract start and end dates for Sanya",

    #     # Auto pay [true]
    #     "I want to verify autopay is set up for account 993871",
    #     "Confirm auto payment is active for Sanya",

    #     # Auto pay [false]
    #     "Auto payment is not set up for account 993871",
    #     "Autopay is disabled, what should I do for Sanya?",

    #     # Auto pay set up
    #     "Set up auto pay for my account 993871",
    #     "Configure automatic bill pay for Sanya",

    #     # Transfer
    #     "Transfer service for account 993871",
    #     "Moving from 7082, need to transfer service",

    #     # Enrollment
    #     "I want to enroll for zip code 7082",
    #     "Same-day enrollment possible at 7082?",

    #     # Power outage [smart meter]
    #     "Power is down, smart meter enabled account 993871",
    #     "No power at 7082, I have a smart meter",

    #     # Power outage [No smart meter]
    #     "Reporting a power outage for account 993871",
    #     "Report power cut for last name Sanya",

    #     # DNP/RNP
    #     "How do I prevent disconnection for account 993871?",
    #     "RNP request for account 993871",

    #     # Cancel service [status Active]
    #     "How do I cancel my active account 993871?",
    #     "Cancel service at zipcode 7082",

    #     # Cancel service [status Terminated]
    #     "Verify cancellation for account 993871",
    #     "My account 993871 is terminated, need refund",

    #     # Cancel service [status Pending]
    #     "Cancel is pending for account 993871, need update",
    #     "Is my pending cancellation going through for Sanya?",
    # ],

    # "bot_5": [
    #     # Pay bill [quick pay]
    #     "I need to make a quick payment for account A-29C6EA19",
    #     "Pay my bill fast using zip 77025",

    #     # Irrelevant chat
    #     "Can you write me a poem?",
    #     "Do you know any fun facts?",

    #     # New customer [General FAQ and Enrollment]
    #     "How quickly can I enroll at zip 77025?",
    #     "I want to switch to Champions Energy in 77025",

    #     # Billing [current balance]
    #     "What's my current bill amount for account A-29C6EA19?",
    #     "What are my current charges for Erlichman?",

    #     # Billing [historic balance]
    #     "What did I pay last month for account A-29C6EA19?",
    #     "Show me old bills for last name Erlichman",

    #     # Current rate plan
    #     "What plan is account A-29C6EA19 on?",
    #     "Current plan information for Erlichman",

    #     # Renewal
    #     "Renewal options for account A-29C6EA19",
    #     "My contract ends soon, how do I renew for Erlichman?",

    #     # Contract details
    #     "What's my early termination fee for account A-29C6EA19?",
    #     "Contract start and end dates for Erlichman",

    #     # Auto pay [true]
    #     "I want to verify autopay is set up for account A-29C6EA19",
    #     "Confirm auto payment is active for Erlichman",

    #     # Auto pay [false]
    #     "Auto payment is not set up for account A-29C6EA19",
    #     "Autopay is disabled, what should I do for Erlichman?",

    #     # Auto pay set up
    #     "Set up auto pay for my account A-29C6EA19",
    #     "Configure automatic bill pay for Erlichman",

    #     # Transfer
    #     "Transfer service for account A-29C6EA19",
    #     "Moving from 77025, need to transfer service",

    #     # Enrollment
    #     "I want to enroll for zip code 77025",
    #     "I need to enroll before my move-in date at 77025",

    #     # Power outage [smart meter]
    #     "Power is down, smart meter enabled account A-29C6EA19",
    #     "No power at 77025, I have a smart meter",

    #     # Power outage [No smart meter]
    #     "Reporting a power outage for account A-29C6EA19",
    #     "Report power cut for last name Erlichman",

    #     # DNP/RNP
    #     "How do I prevent disconnection for account A-29C6EA19?",
    #     "RNP request for account A-29C6EA19",

    #     # Cancel service [status Active]
    #     "How do I cancel my active account A-29C6EA19?",
    #     "Cancel service at zipcode 77025",

    #     # Cancel service [status Terminated]
    #     "Verify cancellation for account A-29C6EA19",
    #     "My account A-29C6EA19 is terminated, need refund",

    #     # Cancel service [status Pending]
    #     "Cancel is pending for account A-29C6EA19, need update",
    #     "Is my pending cancellation going through for Erlichman?",
    # ],
    # "bot_6":  [
    #     "Can I apply for a Zolve account without a PAN card?",
    #     "Do I need to be physically present in the US to activate my account?",
    #     "What should I do if my passport or visa upload fails?",
    #     "What is the typical credit limit range for Zolve users?",
    #     "How long does it take for a physical Zolve card to be delivered?",
    #     "Does Zolve provide loans?",
    #     "Am I eligible to get my loan refinanced?",
    #     "Is there a cosigner requirement for refinancing?",
    #     "Is there a minimum loan amount requirement for refinancing?",
    #     "Can I refinance my Indian provider education loan?",
    #     "Does Zolve refinance loans?",
    #     "i want to talk to real agent"
    #     ],

    # "bot_7":  [
    #     "What steps should I take if my card does not arrive?",
    #     "Is a US phone number required to receive my physical card?",
    #     "Can dependents apply for Zolve accounts?",
    #     "What happens if an international money transfer to Zolve fails?",
    #     "How can I update my phone number?",
    #     "If I apply for refinancing, is it guaranteed that I will receive an option?",
    #     "Why should I buy Health Insurance?",
    #     "What is the University waiver process?",
    #     "Is the waiver process cumbersome?",
    #     "I bought the plan but the university hasn't accepted it.",
    #     "i want to talk to real agent"

    #     ],

    # "bot_8":  [
    #     "How do I update my email ID?",
    #     "i wanted to change my address, can you help me with it?",
    #     "Can my credit card statement reflect my new address immediately?",
    #     "How can I change/reset my password?",
    #     "How do I order a physical card?",
    #     "How do I get refunds and claims?",
    #     "I am not a student. Can I buy these plans?",
    #     "My university is not listed. What should I do?",
    #     "Can you please suggest me a health insurance plan?",
    #     "What is the basic plan, and what does the plan cover?",
    #     "I have an OPT visa type. Can I get health insurance?",
    #     "can u pls connect me to agent?"
    #     ],

    # "bot_9":  [
    #     "How do I update my email ID?",
    #     "How do I update my address?",
    #     "Can my credit card statement reflect my new address immediately?",
    #     "How can I change/reset my password?",
    #     "What are the benefits of the Zolve Classic, Signature, and Black credit cards?",
    #     "Who is eligible to apply for Zolve products?",
    #     "How do I order a physical card?",
    #     "can u pls connect me to agent?"
    #     ],

    # "bot_10": [
    #     "How do I earn rewards?",
    #     "How do I redeem rewards?",
    #     "How can I add money?",
    #     "What are the transfer timelines?",
    #     "Can I pay my credit card bill in cash?",
    #     "What are key billing dates?",
    #     "Does the application do a hard credit check?",
    #     "Can I get a loan for bikes?",
    #     "When is my credit reported?",
    #     "can u pls connect me to agent?"
    #     ],


}

# ── Bot credentials (add/remove bots here) ────────────────────
BOT_CLIENTS = [
    {
        "bot_id":      "bot_1",
        "username":    "guest01fb09c0647c5480b7b09646380863",
        "password":    "723ec21636",
        "recipient":   "support_champion_ticket_demo@xmpp.adjetter.com",
        "task_id":     657661035,
        "support_id":  "support_champion_ticket_demo",
        "support_key": "e03bdaa09c8e11b590e9062a151ee946da3ac94d6179867799",
    },
#     {
#     "bot_id":      "bot_2",
#     "username":    "guest01fb09c0647c5480b7b05285700823",
#     "password":    "68d549512d",
#     "recipient":   "support_champion_ticket_demo@xmpp.adjetter.com",
#     "task_id":     657658871,
#     "support_id":  "support_champion_ticket_demo",
#     "support_key": "e03bdaa09c8e11b590e9062a151ee946da3ac94d6179867799",
# },
# {
#     "bot_id":      "bot_3",
#     "username":    "guest01fb09c0647c5480b7b02172041990",
#     "password":    "f437de9bb4",
#     "recipient":   "support_champion_ticket_demo@xmpp.adjetter.com",
#     "task_id":     657658887,
#     "support_id":  "support_champion_ticket_demo",
#     "support_key": "e03bdaa09c8e11b590e9062a151ee946da3ac94d6179867799",
# },
    # {
    #     "bot_id":      "bot_4",
    #     "username":    "guest01fb09c0647c5480b7b05285700823",
    #     "password":    "68d549512d",
    #     "recipient":   "support_champion_ticket_demo@xmpp.adjetter.com",
    #     "task_id":     657658871,
    #     "support_id":  "support_champion_ticket_demo",
    #     "support_key": "e03bdaa09c8e11b590e9062a151ee946da3ac94d6179867799",
    # },

    # {
    #     "bot_id":      "bot_5",
    #     "username":    "guest01fb09c0647c5480b7b08714457393",
    #     "password":    "634147bf80",
    #     "recipient":   "support_champion_ticket_demo@xmpp.adjetter.com",
    #     "task_id":     657658872,
    #     "support_id":  "support_champion_ticket_demo",
    #     "support_key": "e03bdaa09c8e11b590e9062a151ee946da3ac94d6179867799",

    # },
    # {
    #     "bot_id":      "bot_6",
    #     "username":    "cusa316425101",
    #     "password":    "a65cd29ee6c5df5",
    #     "recipient":   "support_key1007060_ticket_zolvebot@xmpp.adjetter.com",
    #     "task_id":     316425101,
    #     "support_id":  "support_key1007060_ticket_zolvebot",
    #     "support_key": "803c10e26d8e176df07adc7696a33e4863a90a3e1159503618",
    # },
    # {
    #     "bot_id":      "bot_7",
    #     "username":    "cusa316426248",
    #     "password":    "0823f78ec51513e",
    #     "recipient":   "support_key1007060_ticket_zolvebot@xmpp.adjetter.com",
    #     "task_id":     316426248,
    #     "support_id":  "support_key1007060_ticket_zolvebot",
    #     "support_key": "803c10e26d8e176df07adc7696a33e4863a90a3e1159503618",
    # },
    # {
    #     "bot_id":      "bot_8",
    #     "username":    "cusa308298800",
    #     "password":    "3dc6cbc2d6b763f",
    #     "recipient":   "support_key1007060_ticket_zolvebot@xmpp.adjetter.com",
    #     "task_id":     308298800,
    #     "support_id":  "support_key1007060_ticket_zolvebot",
    #     "support_key": "803c10e26d8e176df07adc7696a33e4863a90a3e1159503618",
    # },
    # {
    #     "bot_id":      "bot_9",
    #     "username":    "cusa309215465",
    #     "password":    "d69c9ba4c20a3f8",
    #     "recipient":   "support_key1007060_ticket_zolvebot@xmpp.adjetter.com",
    #     "task_id":     309215465,
    #     "support_id":  "support_key1007060_ticket_zolvebot",
    #     "support_key": "803c10e26d8e176df07adc7696a33e4863a90a3e1159503618",
    # },
    # {
    #     "bot_id":      "bot_10",
    #     "username":    "guest01fb09c0647c5480b7b01781394436",
    #     "password":    "49376319f5",
    #     "recipient":   "support_key1007060_ticket_zolvebot@xmpp.adjetter.com",
    #     "task_id":     1025594290,
    #     "support_id":  "support_key1007060_ticket_zolvebot",
    #     "support_key": "803c10e26d8e176df07adc7696a33e4863a90a3e1159503618",
    # },
]


# ─────────────────────────────────────────────────────────────
#  Excel Logger  (shared singleton with write lock)
# ─────────────────────────────────────────────────────────────
class ExcelLogger:
    _instance = None
    _lock = Semaphore()

    @classmethod
    def get_instance(cls, filename):
        with cls._lock:
            if cls._instance is None:
                cls._instance = cls(filename)
            return cls._instance

    def __init__(self, filename):
        self.filename = filename
        self._write_lock = Semaphore()
        self._load_or_create_workbook()

    def _load_or_create_workbook(self):
        if os.path.exists(self.filename):
            self.wb = load_workbook(self.filename)
            self.ws = self.wb.active
        else:
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.append(["Timestamp", "Bot ID", "Sent Message", "Reply"])
            self.wb.save(self.filename)

    def log(self, bot_id, sent_msg, reply_msg):
        timestamp = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
        print(f"[{timestamp}] [{bot_id}] Sent: {sent_msg} | Reply: {reply_msg}")
        with self._write_lock:
            self.ws.append([timestamp, bot_id, sent_msg, reply_msg])
            self.wb.save(self.filename)


# ─────────────────────────────────────────────────────────────
#  Message utilities
# ─────────────────────────────────────────────────────────────
def extract_message(json_str):
    try:
        decoded = html.unescape(json_str)
        data = json.loads(decoded)
        if "supportChatMessage" in data and "message" in data["supportChatMessage"]:
            msg = data["supportChatMessage"]["message"]
        elif "message" in data:
            msg = data["message"]
        else:
            msg = decoded
        if isinstance(msg, str):
            msg = msg.encode('utf-8').decode('unicode_escape').replace('\n', ' ').strip()
        return msg
    except Exception:
        return json_str


# ─────────────────────────────────────────────────────────────
#  Single-bot WebSocket runner  (no locust, plain gevent)
# ─────────────────────────────────────────────────────────────
class BotRunner:
    def __init__(self, client_cfg):
        self.cfg        = client_cfg
        self.bot_id     = client_cfg["bot_id"]
        self.username   = client_cfg["username"]
        self.password   = client_cfg["password"]
        self.recipient  = client_cfg["recipient"]
        self.logger     = ExcelLogger.get_instance(LOG_FILE)

        # Build payloads for this bot's questions
        self.queries = [
            build_payload(
                msg,
                client_cfg["username"],
                client_cfg["task_id"],
                client_cfg["support_key"],
                client_cfg["support_id"],
            )
            for msg in BOT_QUESTION_SETS.get(self.bot_id, [])
        ]

        self.jid              = ""
        self.connected        = False
        self.awaiting_reply   = False
        self.current_index    = 0
        self.sent_indices     = set()
        self.done             = False          # True when all replies received
        self.send_lock        = Semaphore()
        self._hb_greenlet     = None
        self.ws_app           = None

    # ── Send next queued message ──────────────────────────────
    def _send_next(self, ws):
        with self.send_lock:
            while self.current_index in self.sent_indices:
                self.current_index += 1

            if self.current_index >= len(self.queries):
                print(f"[{self.bot_id}] ✅ All questions sent & answered.")
                self.done = True
                return

            if not self.connected or self.awaiting_reply:
                return

            msg           = self.queries[self.current_index]
            self.last_sent = msg
            new_trace_id  = str(uuid.uuid4())
            current_ms    = int(time.time() * 1000)

            modified = re.sub(r'"traceId"\s*:\s*"[a-fA-F0-9\-]+"',
                               f'"traceId":"{new_trace_id}"', msg)
            modified = re.sub(r'"sentDate":\d+',
                               f'"sentDate":{current_ms}', modified)

            body    = escape(modified)
            msg_id  = f"msg-{uuid.uuid4().int % 10000000000}"
            xmpp    = f"""<message from='{self.jid}' to='{self.recipient}'
                          type='chat' id='{msg_id}' xmlns='jabber:client'>
                            <body>{body}</body>
                          </message>"""
            try:
                ws.send(xmpp)
                self.awaiting_reply = True
                sent_text = extract_message(msg)
                print(f"[{self.bot_id}] → {sent_text}")
            except Exception as e:
                print(f"[{self.bot_id}] ❌ Send failed: {e}")

    # ── Heartbeat ─────────────────────────────────────────────
    def _heartbeat(self, ws):
        try:
            while self.connected:
                try:
                    ws.send("<ping xmlns='urn:xmpp:ping'/>")
                except Exception as e:
                    print(f"[{self.bot_id}] ⚠ Heartbeat error: {e}")
                    break
                gevent.sleep(15)
        finally:
            self._hb_greenlet = None

    # ── on_message handler ────────────────────────────────────
    def _on_message(self, ws, message):
        # Bind success → session ready
        if "<iq" in message and "bind" in message and "type='result'" in message:
            m = re.search(r"<jid>(.*?)</jid>", message)
            self.jid       = m.group(1) if m else "unknown"
            self.connected = True
            self.awaiting_reply = False
            if not self._hb_greenlet:
                self._hb_greenlet = gevent.spawn(self._heartbeat, ws)
            gevent.spawn_later(0.5, self._send_next, ws)
            return

        # Incoming reply
        if "<body>" in message:
            bm = re.search(r"<body>(.*?)</body>", message, re.DOTALL)
            if bm:
                body_content = html.unescape(bm.group(1))
                reply_text   = extract_message(body_content)
                sent_text    = extract_message(self.last_sent)
                self.logger.log(self.bot_id, sent_text, reply_text)

                self.awaiting_reply = False
                self.sent_indices.add(self.current_index)
                self.current_index += 1

                # All done?
                if self.current_index >= len(self.queries):
                    print(f"[{self.bot_id}] ✅ Finished all questions.")
                    self.done = True
                    ws.close()
                    return

                gevent.spawn_later(random.uniform(7, 8), self._send_next, ws)
            return

        # SASL
        if "<stream:features" in message and "PLAIN" in message:
            auth_str = f"\0{self.username}\0{self.password}"
            auth_b64 = base64.b64encode(auth_str.encode()).decode()
            ws.send(f"<auth xmlns='urn:ietf:params:xml:ns:xmpp-sasl' mechanism='PLAIN'>{auth_b64}</auth>")
            return

        if "<success" in message:
            gevent.sleep(1)
            ws.send("<open xmlns='urn:ietf:params:xml:ns:xmpp-framing' to='xmpp.adjetter.com' version='1.0'/>")
            return

        if "<stream:features" in message and "urn:ietf:params:xml:ns:xmpp-bind" in message:
            ws.send("""<iq type='set' id='bind_1' xmlns='jabber:client'>
                          <bind xmlns='urn:ietf:params:xml:ns:xmpp-bind'/>
                       </iq>""")
            return

    # ── on_close handler ─────────────────────────────────────
    def _on_close(self, ws, code, reason):
        self.connected      = False
        self.awaiting_reply = False
        if self._hb_greenlet:
            try:
                self._hb_greenlet.kill()
            except Exception:
                pass
            self._hb_greenlet = None
        print(f"[{self.bot_id}] 🔒 Closed (code={code}): {reason}")

    # ── Main run (blocking inside a greenlet) ─────────────────
    def run(self):
        print(f"[{self.bot_id}] 🚀 Starting …")
        self.ws_app = websocket.WebSocketApp(
            "wss://chat.kapdesk.com/service",
            on_open=lambda ws: ws.send(
                "<open xmlns='urn:ietf:params:xml:ns:xmpp-framing' "
                "to='xmpp.adjetter.com' version='1.0'/>"
            ),
            on_message=self._on_message,
            on_error=lambda ws, err: print(f"[{self.bot_id}] ❌ WS error: {err}"),
            on_close=self._on_close,
        )
        # Blocking — returns when socket is closed (either done or error)
        self.ws_app.run_forever(ping_interval=20, ping_timeout=10)
        print(f"[{self.bot_id}] 🏁 Runner exited (done={self.done})")


# ─────────────────────────────────────────────────────────────
#  Locust User  — spawns all bots in parallel, waits for finish
# ─────────────────────────────────────────────────────────────
class MultiClientUser(User):
    # Locust settings: run once, then stop
    wait_time = lambda self, *_: 1

    @task
    def run_all_bots(self):
        runners   = [BotRunner(cfg) for cfg in BOT_CLIENTS]
        greenlets = [gevent.spawn(r.run) for r in runners]

        # Poll until every bot is done (or greenlet exits)
        while True:
            all_done = all(r.done or g.dead for r, g in zip(runners, greenlets))
            if all_done:
                break
            gevent.sleep(2)

        # Clean up any lingering greenlets
        for g in greenlets:
            if not g.dead:
                g.kill()

        print("✅ All bots finished. Stopping Locust.")
        # Stop the Locust runner gracefully
        self.environment.runner.quit()