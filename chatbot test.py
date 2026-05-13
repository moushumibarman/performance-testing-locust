from locust import User, task
import websocket
import time
import base64
import uuid
import re
import json
import html
from openpyxl import Workbook, load_workbook
import os
from xml.sax.saxutils import escape
import gevent
from gevent.lock import Semaphore
import random
from gevent import monkey

# ----------- GEVENT MONKEY PATCH -----------
monkey.patch_all()

LOG_FILE = "snabbit123.xlsx"

# ----------------- Excel Logger -----------------
class ExcelLogger:
    def __init__(self, filename):
        self.filename = filename
        self._load_or_create_workbook()

    def _load_or_create_workbook(self):
        if os.path.exists(self.filename):
            self.wb = load_workbook(self.filename)
            self.ws = self.wb.active
        else:
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.append(["Timestamp", "User", "Sent Message", "Reply"])
            self.wb.save(self.filename)

    def log(self, user, sent_msg, reply_msg):
        timestamp = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
        print(f"[{timestamp}] [{user}] Sent: {sent_msg} | Reply: {reply_msg}")
        self.ws.append([timestamp, user, sent_msg, reply_msg])
        self.wb.save(self.filename)

# ----------------- Message Utilities -----------------
def extract_message(json_str):
    try:
        decoded = html.unescape(json_str)
        data = json.loads(decoded)
        if "supportChatMessage" in data and "message" in data["supportChatMessage"]:
            msg = data["supportChatMessage"]["message"]
        elif "message" in data:
            msg = data["message"]
        else:
            msg = decoded
        if isinstance(msg, str):
            msg = msg.encode('utf-8').decode('unicode_escape').replace('\n', ' ').strip()
        return msg
    except Exception:
        return json_str

def print_and_log(logger, user, sent_msg, reply_msg=""):
    sent_text = extract_message(sent_msg)
    reply_text = extract_message(reply_msg)
    logger.log(user, sent_text, reply_text)

# ----------------- Base WebSocket User -----------------
class WebSocketUser(User):
    abstract = True

    def on_start(self):
        self.logger = ExcelLogger(LOG_FILE)
        self.jid = ""
        self.queries = []
        self.last_sent_message = ""
        self.ws_app = None
        self.current_index = 0
        self.connected = False
        self.awaiting_reply = False
        self.send_lock = Semaphore()
        self.sent_indices = set()            # track messages we have received replies for
        self._connecting = False             # prevents concurrent connection attempts
        self._runner = None                  # greenlet running run_forever
        self._heartbeat_greenlet = None

    def connect_and_run(self, credentials, recipient, queries):
        self.queries = queries
        user = credentials["username"]

        # ----------------- Send Next Message -----------------
        def send_next(ws):
            # send sequentially, skip indices we already got replies for
            with self.send_lock:
                # advance index if already acknowledged
                while self.current_index in self.sent_indices:
                    self.current_index += 1

                if self.current_index < len(self.queries) and self.connected and not self.awaiting_reply:
                    msg = self.queries[self.current_index]
                    self.last_sent_message = msg
                    message_id = f"msg-{uuid.uuid4().int % 10000000000}"
                    new_trace_id = str(uuid.uuid4())
                    current_ms = int(time.time() * 1000)

                    # Replace traceId and sentDate dynamically
                    modified_msg = re.sub(
                        r'"traceId"\s*:\s*"[a-fA-F0-9\-]+"',
                        f'"traceId":"{new_trace_id}"',
                        msg
                    )
                    modified_msg = re.sub(
                        r'"sentDate":\d+',
                        f'"sentDate":{current_ms}',
                        modified_msg
                    )

                    body = escape(modified_msg)

                    xmpp_message = f"""
                    <message from='{self.jid}'
                            to='{recipient}'
                            type='chat' id='{message_id}'
                            xmlns='jabber:client'>
                        <body>{body}</body>
                    </message>
                    """

                    try:
                        ws.send(xmpp_message)
                        self.awaiting_reply = True
                        print_and_log(self.logger, user, msg)
                    except Exception as e:
                        print(f"[{user}] ❌ Failed to send message: {e}")

        # ----------------- Heartbeat Ping -----------------
        def send_heartbeat(ws):
            try:
                while self.connected:
                    try:
                        ws.send("<ping xmlns='urn:xmpp:ping'/>")
                    except Exception as e:
                        print(f"[{user}] ⚠ Heartbeat failed: {e}")
                        break
                    gevent.sleep(15)
            finally:
                # clear heartbeat reference when it ends
                self._heartbeat_greenlet = None

        # ----------------- Handle Incoming Message -----------------
        def on_message(ws, message):
            # Successful bind (session ready)
            if "<iq" in message and "bind" in message and "type='result'" in message:
                match = re.search(r"<jid>(.*?)</jid>", message)
                self.jid = match.group(1) if match else "unknown"
                self.connected = True
                self.awaiting_reply = False
                # start heartbeat if not running
                if not self._heartbeat_greenlet:
                    self._heartbeat_greenlet = gevent.spawn(send_heartbeat, ws)
                # schedule first send shortly after bind
                gevent.spawn_later(0.5, send_next, ws)
                return

            # Incoming chat message body
            if "<body>" in message:
                body_match = re.search(r"<body>(.*?)</body>", message, re.DOTALL)
                if body_match:
                    body_content = html.unescape(body_match.group(1))
                    reply_text = extract_message(body_content)
                    print_and_log(self.logger, user, self.last_sent_message, reply_text)
                    # mark current index as delivered
                    self.awaiting_reply = False
                    self.sent_indices.add(self.current_index)
                    self.current_index += 1
                    gevent.spawn_later(random.uniform(7, 8), send_next, ws)
                return

            # SASL authentication
            if "<stream:features" in message and "PLAIN" in message:
                auth_str = f"\0{credentials['username']}\0{credentials['password']}"
                auth_b64 = base64.b64encode(auth_str.encode()).decode()
                try:
                    ws.send(f"<auth xmlns='urn:ietf:params:xml:ns:xmpp-sasl' mechanism='PLAIN'>{auth_b64}</auth>")
                except Exception as e:
                    print(f"[{user}] ❌ Auth send failed: {e}")
                return

            if "<success" in message:
                gevent.sleep(1)
                try:
                    ws.send("<open xmlns='urn:ietf:params:xml:ns:xmpp-framing' to='xmpp.adjetter.com' version='1.0'/>")
                except Exception as e:
                    print(f"[{user}] ❌ post-success open failed: {e}")
                return

            if "<stream:features" in message and "urn:ietf:params:xml:ns:xmpp-bind" in message:
                try:
                    ws.send("""
                    <iq type='set' id='bind_1' xmlns='jabber:client'>
                        <bind xmlns='urn:ietf:params:xml:ns:xmpp-bind'/>
                    </iq>
                    """)
                except Exception as e:
                    print(f"[{user}] ❌ bind send failed: {e}")
                return

        # ----------------- Handle Close -----------------
        def on_close(ws, code, reason):
            # normalize code
            try:
                code_int = int(code) if code else None
            except Exception:
                code_int = None

            self.connected = False
            self.awaiting_reply = False
            print(f"[{user}] 🔒 Closed: {code_int} | reason: {reason}")

            # if heartbeat greenlet exists, kill it
            if self._heartbeat_greenlet:
                try:
                    self._heartbeat_greenlet.kill()
                except Exception:
                    pass
                self._heartbeat_greenlet = None

            # Don't attempt another concurrent connect
            self._connecting = False
            # Normal closure -> wait longer with jitter before reconnecting to avoid flood
            if code_int == 1000:
                wait = random.uniform(30, 90)  # 30-90s wait on normal close
                print(f"[{user}] Normal closure (1000). Will reconnect after {wait:.1f}s to avoid flood.")
                gevent.sleep(wait)
                # continue to reconnect after long wait
                if not self._runner or self._runner.dead:
                    self._runner = gevent.spawn(run_forever)
            else:
                # abnormal closure -> retry sooner but with exponential backoff & jitter
                # note: run_forever has its own backoff loop, so simply ensure it's running
                backoff = random.uniform(2, 6)
                print(f"[{user}] Abnormal closure. Will attempt reconnect in {backoff:.1f}s")
                gevent.sleep(backoff)
                if not self._runner or self._runner.dead:
                    self._runner = gevent.spawn(run_forever)

        # ----------------- Run Forever with guarded reconnect -----------------
        def run_forever():
            if self._connecting:
                # another runner is already connecting
                return
            self._connecting = True
            backoff = 5
            try:
                while True:
                    try:
                        # create a fresh WebSocketApp for each attempt
                        self.ws_app = websocket.WebSocketApp(
                            "wss://chat.kapdesk.com/service",
                            on_open=lambda ws: ws.send("<open xmlns='urn:ietf:params:xml:ns:xmpp-framing' to='xmpp.adjetter.com' version='1.0'/>"),
                            on_message=on_message,
                            on_error=lambda ws, err: print(f"[{user}] ❌ Error: {err}"),
                            on_close=on_close
                        )
                        # blocking call - returns when closed
                        self.ws_app.run_forever(ping_interval=20, ping_timeout=10)
                        # when run_forever returns, the socket is closed - mark disconnected
                        self.connected = False
                    except Exception as e:
                        print(f"[{user}] ⚠ WebSocket error in run_forever: {e}")

                    # if we reach here, the socket was closed; choose a backoff before next attempt
                    sleep_time = backoff + random.uniform(0, 2)
                    print(f"[{user}] run_forever ended — sleeping {sleep_time:.1f}s before next attempt")
                    gevent.sleep(sleep_time)
                    backoff = min(backoff * 2, 60)
            finally:
                self._connecting = False

        # start runner only if not already running
        if not self._runner or self._runner.dead:
            self._runner = gevent.spawn(run_forever)
        # give it a moment to connect
        gevent.sleep(6)


# ----------------- Multi Client User -----------------
class MultiClientUser(WebSocketUser):
    plain_messages = [
        [

            "How can I make a booking?",
            "Is Snabbit available in my area?",
            "What all is included in a service?",
            "Need service for a commercial property",
            "I want service from my favourite expert",
            "Can't see available slots",
            "Cannot apply coupon code",
            "How to use Starter Pack",
            "How to earn referral rewards",
            "Looking for latest offers",
            "Issue with work quality",
            "Faced a behavior issue with the expert",
            "Need to report damage or missing items",
            "Change mode of payment",
            "Why can't I see any available slots?",
            "My expert has not been assigned yet.",
            "My expert hasn't arrived yet.",
            "How can I reschedule a booking?",
            "How can I cancel a booking?"

]

    ]

    queries_array = []
    for msg in plain_messages:
        payload = {
            "key": str(uuid.uuid4().hex),
            "type": "SUPPORT",
            "supportChatMessage": {
                "fromJid": "cusa331091309",
                "senderName": "Snabbit support",
                "message": msg,
                "taskId": 993536148,
                "ticketId": "4770969682632",
                "enquiryId": 0,
                "leadId": 331091309,
                "sentDate": int(time.time() * 1000),
                "supportId": "support_key1007028_ticket_snabbits",
                "supportKey": "3c76670a710b13721ab3fefb7a55619c362b2ccb4412748035",
                "chatbotMessageKey": None,
                "action": None,
                "additionalData": None,
                "productDetail": None,
                "chatbotFlow": None,
                "conversationId": None,
                "isSentByBot": True,
                "traceId": str(uuid.uuid4())
            },
            "traceId": str(uuid.uuid4())
        }
        queries_array.append(json.dumps(payload))

    clients = [
        {
            "username": "cusa331091309",
            "password": "28cc5c85a6ffb44",
            "recipient": "support_key1007028_ticket_snabbits@xmpp.adjetter.com",
            "queries": queries_array
        }
    ]

    connection_lock = Semaphore()

    @task
    def connect_all_clients(self):
        greenlets = []
        for client in self.clients:
            g = gevent.spawn(self._connect_single_client, client)
            greenlets.append(g)
        gevent.joinall(greenlets)

    def _connect_single_client(self, client):
        with self.connection_lock:
            self.logger = ExcelLogger(LOG_FILE)
        self.connect_and_run(
            credentials={"username": client["username"], "password": client["password"]},
            recipient=client["recipient"],
            queries=client["queries"]
        )
